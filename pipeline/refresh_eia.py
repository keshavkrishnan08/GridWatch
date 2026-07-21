#!/usr/bin/env python3
"""
Autonomously refresh the mechanical utility figures from EIA-861.

    python3 -m pipeline.refresh_eia            # report what would change
    python3 -m pipeline.refresh_eia --write    # apply it

EIA Form 861 is published annually as a keyless ZIP at a stable URL, with one
row per utility per state. That makes customer counts and class revenue shares
genuinely automatable — and they're exactly the figures that go quietly stale
and produce wrong per-household math.

WHAT THIS TOUCHES, AND WHAT IT WILL NOT
---------------------------------------
It writes only these fields, and only for utilities it matched confidently:

    customers                       residential customer count
    residential_revenue_share_pct   residential share of retail revenue
    eia_avg_rate_cents_kwh          cross-check against the tariff rate

Everything else is human judgment read out of filings — cost_shifts,
recent_increase, tariff rates, and the notes. An automated job must never
overwrite those, so it doesn't: the allowlist is enforced, not just intended.

SAFETY
------
- A failed fetch changes nothing.
- A utility that matches zero or several EIA rows is skipped and reported.
- A value that moves more than --max-change is flagged for review instead of
  applied, because a big jump usually means the name matching broke, not that
  half a million people moved.
"""

from __future__ import annotations

import argparse
import io
import json
import os
import sys
import zipfile
from datetime import date

import requests

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "public", "data")
BILL = os.path.join(DATA, "bill_impact_models.json")
EIA_ZIP = "https://www.eia.gov/electricity/data/eia861/zip/f861{year}.zip"
UA = {"User-Agent": "GridWatch/1.0 (open-source civic grid atlas)"}

# The only keys this script is permitted to write.
ALLOWED = {"customers", "residential_revenue_share_pct", "eia_avg_rate_cents_kwh"}


def load_sales(year: str, state: str):
    """Return {utility_name_lower: {...class figures...}} from EIA-861."""
    r = requests.get(EIA_ZIP.format(year=year), headers=UA, timeout=300)
    r.raise_for_status()
    z = zipfile.ZipFile(io.BytesIO(r.content))
    target = next(n for n in z.namelist() if n.lower().startswith("sales_ult_cust")
                  and "_cs_" not in n.lower())

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(z.read(target)), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    # Header spans three rows: class names, then metric names, then units.
    # Walk them to build column indices rather than hardcoding positions, so an
    # EIA layout change is survivable.
    group_row = next(rows)
    metric_row = next(rows)
    label_row = next(rows)

    groups: list[str] = []
    cur = ""
    for g in group_row:
        cur = (str(g).strip().upper() if g else cur)
        groups.append(cur)

    cols: dict[tuple[str, str], int] = {}
    for i, m in enumerate(metric_row):
        if m and i < len(groups):
            cols[(groups[i], str(m).strip().lower())] = i
    base = {str(v).strip().lower(): i for i, v in enumerate(label_row) if v}

    def idx(*names):
        for n in names:
            if n in base:
                return base[n]
        return None

    i_name, i_state = idx("utility name"), idx("state")
    i_part, i_svc = idx("part"), idx("service type")
    if i_name is None or i_state is None:
        raise RuntimeError("EIA-861 layout changed: could not find utility/state columns")

    def cell(row, key):
        i = cols.get(key)
        if i is None or i >= len(row):
            return None
        v = row[i]
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    out: dict[str, dict] = {}
    for row in rows:
        if not row or len(row) <= i_state:
            continue
        if str(row[i_state]).strip().upper() != state:
            continue
        # Bundled service, Part A/B = the standard retail rows
        if i_svc is not None and str(row[i_svc]).strip().lower() not in ("bundled", "energy", "delivery"):
            continue
        name = str(row[i_name]).strip()
        rec = out.setdefault(name.lower(), {"name": name, "res_rev": 0.0, "res_mwh": 0.0,
                                            "res_cust": 0.0, "total_rev": 0.0})
        rec["res_rev"] += cell(row, ("RESIDENTIAL", "revenues")) or 0
        rec["res_mwh"] += cell(row, ("RESIDENTIAL", "sales")) or 0
        rec["res_cust"] += cell(row, ("RESIDENTIAL", "customers")) or 0
        for grp in ("RESIDENTIAL", "COMMERCIAL", "INDUSTRIAL", "TRANSPORTATION"):
            rec["total_rev"] += cell(row, (grp, "revenues")) or 0
    return out


def match(util: dict, sales: dict) -> list[dict]:
    """Find EIA rows for a configured utility using its raw_match aliases."""
    aliases = [a.lower() for a in util.get("raw_match", [])] + [util["display_name"].lower()]
    hits = []
    for key, rec in sales.items():
        if any(a and a in key for a in aliases):
            hits.append(rec)
    return hits


def main() -> int:
    ap = argparse.ArgumentParser(description="Refresh mechanical utility figures from EIA-861.")
    ap.add_argument("--year", default="2024")
    ap.add_argument("--state", default="IN")
    ap.add_argument("--write", action="store_true", help="apply changes (default: dry run)")
    ap.add_argument("--max-change", type=float, default=25.0,
                    help="flag rather than apply a change larger than this %% (default 25)")
    args = ap.parse_args()

    try:
        with open(BILL) as fh:
            bill = json.load(fh)
    except Exception as e:
        print(f"  ERROR  could not read {BILL}: {e}")
        return 2

    print(f"\n  EIA-861 {args.year} · {args.state} · {'WRITE' if args.write else 'dry run'}")
    try:
        sales = load_sales(args.year, args.state)
    except Exception as e:
        print(f"  ERROR  fetch/parse failed: {e}\n  Nothing changed — re-run to retry.\n")
        return 2
    print(f"  {len(sales)} utilities in the EIA file\n" + "-" * 64)

    changes, flagged, skipped = [], [], []
    for u in bill.get("utilities", []):
        hits = match(u, sales)
        if len(hits) != 1:
            skipped.append((u["display_name"], f"{len(hits)} EIA matches"))
            continue
        rec = hits[0]
        if rec["res_cust"] <= 0 or rec["total_rev"] <= 0:
            skipped.append((u["display_name"], "incomplete EIA row"))
            continue

        proposed = {
            "customers": int(round(rec["res_cust"])),
            "residential_revenue_share_pct": round(100 * rec["res_rev"] / rec["total_rev"], 1),
            # revenues are thousands of dollars, sales are MWh -> cents/kWh
            "eia_avg_rate_cents_kwh": round((rec["res_rev"] * 1000) / (rec["res_mwh"] * 1000) * 100, 2)
            if rec["res_mwh"] > 0 else None,
        }

        for k, new in proposed.items():
            if new is None or k not in ALLOWED:
                continue
            old = u.get(k)
            if old == new:
                continue
            if isinstance(old, (int, float)) and old:
                delta = abs(new - old) / abs(old) * 100
                if delta > args.max_change:
                    flagged.append((u["display_name"], k, old, new, delta))
                    continue
            changes.append((u["display_name"], k, old, new))
            if args.write:
                u[k] = new

    for name, k, old, new in changes:
        print(f"  update   {name:26} {k:30} {old} -> {new}")
    for name, k, old, new, d in flagged:
        print(f"  FLAG     {name:26} {k:30} {old} -> {new}  ({d:.0f}% change — review by hand)")
    for name, why in skipped:
        print(f"  skip     {name:26} {why}")

    print("-" * 64)
    print(f"  {len(changes)} updates, {len(flagged)} flagged, {len(skipped)} skipped")

    if args.write and changes:
        bill.setdefault("_auto", {})
        bill["_auto"]["eia_861"] = {
            "year": args.year,
            "refreshed": date.today().isoformat(),
            "fields": sorted(ALLOWED),
            "note": "Refreshed by pipeline/refresh_eia.py. Only these fields are "
                    "machine-maintained; cost_shifts, recent_increase, tariff rates "
                    "and notes are human-verified from filings and are never touched.",
            "source": f"https://www.eia.gov/electricity/data/eia861/ (Form 861, {args.year})",
        }
        with open(BILL, "w") as fh:
            json.dump(bill, fh, indent=2)
            fh.write("\n")
        print(f"  written -> {os.path.relpath(BILL, ROOT)}")
    elif changes:
        print("  dry run — re-run with --write to apply")
    print()
    return 1 if flagged else 0


if __name__ == "__main__":
    sys.exit(main())
